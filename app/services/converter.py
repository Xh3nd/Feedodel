from io import BytesIO
from typing import Protocol

from openpyxl import load_workbook
from lxml import etree


class PlatformConverter(Protocol):
    def build_xml(self, wb) -> etree._Element:
        ...


def _default_platform_converter(wb) -> etree._Element:
    """
    Very simple default XML:

    <Workbook>
      <Sheet name="Sheet1">
        <Row index="1">
          <Cell col="A">value</Cell>
          ...
        </Row>
      </Sheet>
    </Workbook>
    """
    root = etree.Element("Workbook")

    for sheet in wb.worksheets:
        sheet_el = etree.SubElement(root, "Sheet", name=sheet.title)
        for row in sheet.iter_rows(values_only=True):
            # openpyxl rows are 1-based when using row=row[0].row, so we just track index ourselves
            row_index = row[0].row if hasattr(row[0], "row") else None  # type: ignore[attr-defined]

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_el = etree.SubElement(sheet_el, "Row", index=str(i))
            for j, value in enumerate(row, start=1):
                col_name = chr(ord("A") + j - 1)
                cell_el = etree.SubElement(row_el, "Cell", col=col_name)
                if value is not None:
                    cell_el.text = str(value)

    return root


PLATFORM_CONVERTERS: dict[str, PlatformConverter] = {
    "default": _default_platform_converter,
    # future: add "platform_x": converter_function
}


def convert_excel_to_xml(content: bytes, platform: str = "default") -> bytes:
    if platform not in PLATFORM_CONVERTERS:
        raise ValueError(f"Unsupported platform '{platform}'.")

    wb = load_workbook(filename=BytesIO(content), data_only=True)

    builder = PLATFORM_CONVERTERS[platform]
    root = builder(wb)

    xml_bytes = etree.tostring(
        root,
        xml_declaration=True,
        encoding="utf-8",
        pretty_print=True,
    )
    return xml_bytes


